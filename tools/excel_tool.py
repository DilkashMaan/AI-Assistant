

import subprocess
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config


# ── Styling constants ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2E4057")   # Dark navy
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ROW_FILL_A  = PatternFill("solid", fgColor="F5F8FA")   # Light grey
ROW_FILL_B  = PatternFill("solid", fgColor="FFFFFF")   # White
DATA_FONT   = Font(name="Calibri", size=10)
BORDER_SIDE = Side(style="thin", color="D0D7DE")
CELL_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)


def create_workbook(
    headers: list[str],
    rows: list[list[str]],
    csv_path: Path,
    sheet_title: str = "Data",
) -> Path:
    
    xlsx_path = csv_path.with_suffix(".xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  

    # ── Header row ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = CELL_BORDER

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, start=2):
        fill = ROW_FILL_A if row_idx % 2 == 0 else ROW_FILL_B
        ws.row_dimensions[row_idx].height = 18
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_coerce_value(value))
            cell.font      = DATA_FONT
            cell.fill      = fill
            cell.alignment = LEFT
            cell.border    = CELL_BORDER

    # ── Auto-fit column widths ────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row_data in rows:
            if col_idx - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        # Clamp between 10 and 40
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 4, 40))

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx_path)
    return xlsx_path


def open_in_excel(xlsx_path: Path) -> None:
    """
    Open the xlsx file in Microsoft Excel.
    Tries known Excel paths; falls back to os.startfile on Windows.
    """
    # Try known Excel executable paths first
    for excel_exe in config.EXCEL_PATHS:
        if os.path.isfile(excel_exe):
            subprocess.Popen([excel_exe, str(xlsx_path)])
            return

    # Fall back: let Windows open the file with the default application
    os.startfile(str(xlsx_path))


# ── Helper ─────────────────────────────────────────────────────────────────────
def _coerce_value(value: str):
    """Try to convert a string cell value to int or float; keep as str otherwise."""
    try:
        int_val = int(value)
        return int_val
    except (ValueError, TypeError):
        pass
    try:
        float_val = float(value)
        return float_val
    except (ValueError, TypeError):
        pass
    return value
